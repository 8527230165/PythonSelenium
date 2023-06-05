from selenium import webdriver
import openpyxl
import os


print(os.getcwd())

file=os.getcwd()+"\\Users.xlsx"
workbook=openpyxl.load_workbook(file)

sheet=workbook["Sheet1"]

row=sheet.max_row
column=sheet.max_column

for i in range(1,row+1):
    for j in range(1,column+1):
        print(sheet.cell(i,j).value,end='       ')
    print()